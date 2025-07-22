from flask import Flask, render_template, request, redirect, url_for, send_from_directory, abort, flash, jsonify, session, Response, send_file, current_app
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import os, io, csv, uuid, random
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime
import openpyxl
from io import BytesIO
from openpyxl import load_workbook, Workbook
import random
import string
import secrets
from werkzeug.middleware.proxy_fix import ProxyFix

# Flask и БД
app = Flask(__name__)
app.config['SECRET_KEY'] = 'mysecretkey'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///users.db'
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1)

# Папки для загрузок
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
AVATAR_FOLDER = os.path.join(UPLOAD_FOLDER, 'avatars')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(AVATAR_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['AVATAR_FOLDER'] = AVATAR_FOLDER

# Допустимые форматы
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
MAX_AVATAR_SIZE = 5 * 1024 * 1024  # 5 MB
ICONS = ['📁', '📂', '🗂️', '🗃️', '🧷', '🧮', '📌', '🔖', '🗄️']




# База и логин
db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'auth'

# Вспомогательные
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def set_toast_and_redirect(message, category, endpoint):
    response = redirect(url_for(endpoint))
    response.set_cookie('toast_message', message, max_age=3)
    response.set_cookie('toast_category', category, max_age=3)
    return response

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    temp_password = db.Column(db.String(255), nullable=True)
    password_changed = db.Column(db.Boolean, default=False)
    role = db.Column(db.String(20), default='new')
    first_name = db.Column(db.String(100))
    last_name = db.Column(db.String(100))
    birth_date = db.Column(db.Date, nullable=True)
    phone = db.Column(db.String(20), nullable=True)
    email = db.Column(db.String(120), nullable=True)
    avatar = db.Column(db.String(255), default='default.png')  # только имя файла
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_blocked = db.Column(db.Boolean, default=False)

    @property
    def is_admin(self):
        return self.role == 'admin'

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
        self.temp_password = None
        self.password_changed = True

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class Facility(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    description = db.Column(db.Text, nullable=True)
    type = db.Column(db.String(50), nullable=False)
    district = db.Column(db.String(50), nullable=False)
    documents = db.relationship('Document', backref='facility', lazy=True)
    file_groups = db.relationship("FileGroup", backref="facility", lazy=True)

class FileGroup(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    facility_id = db.Column(db.Integer, db.ForeignKey('facility.id'), nullable=False)
    parent_id = db.Column(db.Integer, db.ForeignKey('file_group.id'), nullable=True)
    children = db.relationship('FileGroup', backref=db.backref('parent', remote_side=[id]), lazy=True)
    documents = db.relationship('Document', backref='group', lazy=True)
    icon = db.Column(db.String(10), default='📁')

class Document(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    facility_id = db.Column(db.Integer, db.ForeignKey('facility.id'), nullable=False)
    group_id = db.Column(db.Integer, db.ForeignKey('file_group.id'), nullable=True)
    name = db.Column(db.String(150), nullable=False)
    description = db.Column(db.Text, nullable=True)
    file_path = db.Column(db.String(255), nullable=False)
    original_name = db.Column(db.String(255), nullable=False)
    file_type = db.Column(db.String(20), nullable=True)

@app.route('/home')
@login_required
def home():
    return render_template('home.html')

# Отдача аватаров
@app.route('/avatars/<filename>')
def avatar(filename):
    path = os.path.join(app.config['AVATAR_FOLDER'], filename)
    if not os.path.isfile(path):
        filename = 'default.png'
    return send_from_directory(app.config['AVATAR_FOLDER'], filename)

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        birth_date_str = request.form.get('birth_date', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email')
        email = email.strip() if email else None

        # Валидация email
        if email and '@' not in email:
            flash('Некорректный email.', 'danger')
            return redirect(url_for('profile'))

        if not first_name or not last_name or not phone:
            flash('Имя, фамилия и телефон обязательны.', 'danger')
            return redirect(url_for('profile'))

        # Дата рождения
        birth_date = None
        if birth_date_str:
            try:
                birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Неверный формат даты рождения.', 'danger')
                return redirect(url_for('profile'))

        # Загрузка аватара
        avatar_file = request.files.get('avatar')
        if avatar_file and avatar_file.filename != '' and allowed_file(avatar_file.filename):
            filename = secure_filename(f"user_{current_user.id}_" + avatar_file.filename)
            avatar_path = os.path.join(app.config['AVATAR_FOLDER'], filename)
            avatar_file.save(avatar_path)

            # Удаление старого аватара (если не default.png)
            if current_user.avatar and current_user.avatar != 'default.png':
                old_avatar_path = os.path.join(app.config['AVATAR_FOLDER'], current_user.avatar)
                if os.path.exists(old_avatar_path):
                    os.remove(old_avatar_path)

            current_user.avatar = filename

        # Обновление данных
        current_user.first_name = first_name
        current_user.last_name = last_name
        current_user.birth_date = birth_date
        current_user.phone = phone
        current_user.email = email

        db.session.commit()
        flash('Профиль успешно обновлён!', 'success')
        return redirect(url_for('profile'))

    return render_template('profile.html')


@app.route('/documents')
@login_required  # если страница только для авторизованных
def documents():
    return render_template('documents.html')

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/contacts')
def contacts():
    is_authenticated = 'user_id' in session  # или current_user.is_authenticated, если используешь flask-login
    return render_template('contacts.html', is_authenticated=is_authenticated)

@app.route('/emergency')
@login_required
def emergency():
    return render_template('emergency.html')

@app.route('/auth', methods=['GET', 'POST'])
def auth():
    if current_user.is_authenticated:
        return redirect(url_for('facilities'))

    error = None
    if request.method == 'POST':
        action = request.form.get('action')
        username = request.form.get('username')
        password = request.form.get('password')

        if not username or not password:
            error = "Пожалуйста, заполните все поля."
        else:
            user = User.query.filter_by(username=username).first()

            if user and user.is_blocked:
                response = redirect(url_for('auth'))
                response.set_cookie('toast_message', "Ваш аккаунт заблокирован.", max_age=3)
                response.set_cookie('toast_category', "danger", max_age=3)
                return response

            if action == 'login':
                if user and user.check_password(password):
                    if user.role == 'new':
                        error = "Ваш аккаунт ещё не подтверждён администратором."
                    else:
                        login_user(user)

                        # Если пользователь не менял пароль — редирект на смену пароля
                        if not user.password_changed:
                            return redirect(url_for('change_password'))

                        return redirect(url_for('facilities'))
                else:
                    error = "Неверное имя пользователя или пароль."


            elif action == 'register':
                existing_user = User.query.filter_by(username=username).first()
                first_name = request.form.get('first_name')
                last_name = request.form.get('last_name')
                if existing_user:
                    error = "Пользователь с таким именем уже существует."
                else:
                    new_user = User(
                        username=username,
                        first_name=first_name,
                        last_name=last_name
                    )
                    new_user.set_password(password)
                    new_user.role = 'new'
                    db.session.add(new_user)
                    db.session.commit()
                    error = "Регистрация успешна! Ожидайте подтверждения администратора."
            else:
                error = "Неверное действие."

    return render_template('auth.html', error=error)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Доступ запрещен. Только для администратора.')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/facilities')
@login_required
def facilities():
    facilities = Facility.query.all()
    return render_template('facilities.html', facilities=facilities)

@app.route('/facilities/add', methods=['GET', 'POST'])
@login_required
def add_facility():
    types = ['Котельная', 'Насосная', 'ЦТП', 'ИТП']
    districts = ['Северный', 'Ленинский', 'Октябрьский', 'Юго-западный']
    if request.method == 'POST':
        name = request.form['name']
        description = request.form['description']
        type = request.form['type']
        district = request.form['district']
        new_fac = Facility(name=name, description=description, type=type, district=district)
        db.session.add(new_fac)
        db.session.commit()
        return redirect(url_for('facilities'))
    return render_template('add_facility.html', types=types, districts=districts)

@app.route('/facility/<int:facility_id>')
@login_required
def facility_detail(facility_id):
    facility = Facility.query.get_or_404(facility_id)
    return render_template('facility_detail.html', facility=facility)

@app.route('/admin/users/import', methods=['POST'])
@admin_required
def import_users():
    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        flash('Пожалуйста, выберите Excel-файл (.xlsx)', 'danger')
        return redirect(url_for('admin_dashboard'))

    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        first_name, last_name, username, password, role = row[:5]

        if not username or not role:
            continue  # обязательные поля

        existing = User.query.filter_by(username=username).first()
        if existing:
            continue

        # Если пароль есть — хэшируем, иначе создаём временный
        if password:
            hashed = generate_password_hash(password)
            temp_password = None
        else:
            temp_password = secrets.token_urlsafe(8)
            hashed = generate_password_hash(temp_password)

        user = User(
            first_name=first_name or "",
            last_name=last_name or "",
            username=username,
            password_hash=hashed,
            temp_password=temp_password,
            role=role,
        )

        db.session.add(user)
        count += 1

    db.session.commit()
    flash(f'Импортировано пользователей: {count}', 'success')
    return redirect(url_for('admin_dashboard'))



@app.route('/admin/export_users')
@admin_required
def export_users():
    users = User.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"

    # Заголовки
    ws.append(["Имя", "Фамилия", "Логин", "Пароль (если есть)", "Роль", "Заблокирован?"])

    for user in users:
        ws.append([
            user.first_name,
            user.last_name,
            user.username,
            user.temp_password or "",  # показываем только если есть
            user.role,
            "Да" if user.is_blocked else "Нет"
        ])

    # Сохраняем в байтовый поток
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='users_export.xlsx'
    )



@app.route('/facility/<int:facility_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_facility(facility_id):
    facility = Facility.query.get_or_404(facility_id)
    types = ['Котельная', 'Насосная', 'ЦТП', 'ИТП']
    districts = ['Северный', 'Ленинский', 'Октябрьский', 'Юго-западный']
    if request.method == 'POST':
        facility.name = request.form['name']
        facility.description = request.form['description']
        facility.type = request.form['type']
        facility.district = request.form['district']
        db.session.commit()
        return redirect(url_for('facility_detail', facility_id=facility.id))
    return render_template('edit_facility.html', facility=facility, types=types, districts=districts)

@app.route('/file/<int:file_id>/delete', methods=['POST'])
@login_required
def delete_file(file_id):
    doc = Document.query.get_or_404(file_id)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], doc.file_path)
    if os.path.exists(filepath):
        os.remove(filepath)
    db.session.delete(doc)
    db.session.commit()
    return redirect(url_for('facility_detail', facility_id=doc.facility_id))

@app.route('/file_group/<int:group_id>/delete', methods=['POST'])
@login_required
def delete_file_group(group_id):
    group = FileGroup.query.get_or_404(group_id)
    facility_id = group.facility_id
    # Опционально: удалять все файлы этой группы или оставить без группы
    for doc in group.documents:
        doc.group_id = None
    db.session.delete(group)
    db.session.commit()
    flash('Группа файлов удалена')
    return redirect(url_for('facility_detail', facility_id=facility_id))

@app.route('/file/<int:file_id>/change_group', methods=['POST'])
@login_required
def change_file_group(file_id):
    new_group_id = request.form.get('group_id')
    file = Document.query.get_or_404(file_id)
    file.group_id = new_group_id if new_group_id else None
    db.session.commit()
    flash('Группа файла изменена')
    return redirect(request.referrer or url_for('facility_detail', facility_id=file.facility_id))

@app.route('/facility/<int:facility_id>/upload', methods=['GET', 'POST'])
@login_required
def upload_file(facility_id):
    facility = Facility.query.get_or_404(facility_id)
    groups = FileGroup.query.filter_by(facility_id=facility.id).all()
    if request.method == 'POST':
        name = request.form['name']
        description = request.form.get('description')
        file_type = request.form['file_type']
        group_id = request.form.get('group_id') or None
        file = request.files['file']
        if file:
            ext = os.path.splitext(file.filename)[1]
            uid = uuid.uuid4().hex
            filename = f"{uid}{ext}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            new_doc = Document(
                name=name,
                description=description,
                file_path=filename,
                original_name=file.filename,
                file_type=file_type,
                facility_id=facility.id,
                group_id=group_id
            )
            db.session.add(new_doc)
            db.session.commit()
            return redirect(url_for('facility_detail', facility_id=facility.id))
    return render_template('upload_file.html', facility=facility, groups=groups)

@app.route('/uploads/<filename>')
@login_required
def download_file(filename):
    doc = Document.query.filter_by(file_path=filename).first()
    if not doc:
        abort(404)
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=False, download_name=doc.original_name)

@app.route('/admin/dashboard', methods=['GET', 'POST'])
@admin_required
def admin_dashboard():
    users = User.query.all()
    # Статистика
    total_users = len(users)
    admin_count = len([u for u in users if u.role == 'admin'])
    pending_count = len([u for u in users if u.role == 'new'])
    today_count = 0  # если нужно, посчитай по created_at

    if request.method == 'POST':
        action = request.form.get('action')
        if not action:
            flash("Действие не указано", "danger")
            return redirect(url_for('admin_dashboard'))

        if action.startswith('save_'):
            user_id = int(action.split('_')[1])
            user = User.query.get(user_id)
            if user and user.id != current_user.id:
                prefix = f'users[{user_id}]'
                user.first_name = request.form.get(f'{prefix}[first_name]', user.first_name)
                user.last_name = request.form.get(f'{prefix}[last_name]', user.last_name)
                user.username = request.form.get(f'{prefix}[username]', user.username)
                user.role = request.form.get(f'{prefix}[role]', user.role)
                is_blocked_value = request.form.get(f'{prefix}[is_blocked]')
                user.is_blocked = bool(is_blocked_value)

                password = request.form.get(f'{prefix}[password]', '').strip()
                if password:
                    if user.temp_password and password == user.temp_password:
                        pass  # не менять, если это тот же временный
                    else:
                        user.set_password(password)
                        user.temp_password = None  # сбрасываем временный
                        user.must_change_password = False

                db.session.commit()
                flash(f"Пользователь {user.username} обновлён.", "success")
                return redirect(url_for('admin_dashboard'))
            else:
                flash("Нельзя редактировать свою учетную запись здесь.", "warning")
                return redirect(url_for('admin_dashboard'))

        elif action.startswith('delete_'):
            user_id = int(action.split('_')[1])
            user = User.query.get(user_id)
            if user and user.id != current_user.id:
                db.session.delete(user)
                db.session.commit()
                flash(f"Пользователь {user.username} удалён.", "danger")
                return redirect(url_for('admin_dashboard'))
            else:
                flash("Нельзя удалить себя.", "warning")
                return redirect(url_for('admin_dashboard'))

    return render_template('admin_dashboard.html',
                           users=users,
                           total_users=total_users,
                           admin_count=admin_count,
                           pending_count=pending_count,
                           today_count=today_count)

@app.route('/admin/users/add', methods=['GET', 'POST'])
@admin_required
def add_user():
    if request.method == 'POST':
        username = request.form['username']
        first_name = request.form['first_name']
        last_name = request.form['last_name']
        password = request.form['password']
        role = request.form['role']
        if User.query.filter_by(username=username).first():
            return redirect(url_for('admin_dashboard', message="Пользователь с таким логином уже существует.", category="warning"))
        new_user = User(
            username=username,
            first_name=first_name,
            last_name=last_name,
            role=role
        )
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        response = redirect(url_for('admin_dashboard'))
        flash(f"Пользователь {username} добавлен.", "success")
        return redirect(url_for('admin_dashboard'))
    return render_template('add_user.html')

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    error = None
    if request.method == 'POST':
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        if not new_password or not confirm_password:
            error = "Пожалуйста, заполните все поля."
        elif new_password != confirm_password:
            error = "Пароли не совпадают."
        else:
            current_user.set_password(new_password)
            db.session.commit()
            flash('Пароль успешно изменён.', 'success')
            return redirect(url_for('facilities'))
    return render_template('change_password.html', error=error)

@app.route('/admin/template.xlsx')
@admin_required
def download_user_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    ws.append(['ФИО', 'Логин', 'Пароль', 'Роль'])  # Пример: Иванов Иван | ivanov | 123 | user

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(stream, download_name="user_template.xlsx", as_attachment=True)

@app.route('/facility/<int:facility_id>/delete', methods=['POST'])
@login_required
def delete_facility(facility_id):
    facility = Facility.query.get_or_404(facility_id)
    db.session.delete(facility)
    db.session.commit()
    flash(f'Объект "{facility.name}" был удалён.')
    return redirect(url_for('facilities'))

@app.route('/facility/<int:facility_id>/groups/add', methods=['GET', 'POST'])
@login_required
def add_file_group(facility_id):
    facility = Facility.query.get_or_404(facility_id)
    if request.method == 'POST':
        group_name = request.form.get('group_name')
        if group_name:
            icon = random.choice(ICONS)  # если у тебя есть список ICONS
            new_group = FileGroup(name=group_name, facility_id=facility_id, icon=icon)
            db.session.add(new_group)
            db.session.commit()
            flash("Группа файлов добавлена.")
            return redirect(url_for('facility_detail', facility_id=facility_id))
        else:
            flash("Название группы обязательно.")
    return render_template('add_file_group.html', facility=facility)

@app.route('/facility/<int:facility_id>/groups/<int:parent_group_id>/add_subgroup', methods=['GET', 'POST'])
@login_required
def add_subgroup(facility_id, parent_group_id):
    facility = Facility.query.get_or_404(facility_id)
    parent_group = FileGroup.query.get_or_404(parent_group_id)
    if request.method == 'POST':
        name = request.form.get('group_name')
        if name:
            icon = random.choice(ICONS)
            subgroup = FileGroup(
                name=name,
                facility_id=facility_id,
                parent_id=parent_group_id,
                icon=icon
            )
            db.session.add(subgroup)
            db.session.commit()
            flash("Подгруппа добавлена.")
            return redirect(url_for('facility_detail', facility_id=facility_id))
        else:
            flash("Название подгруппы обязательно.")
    return render_template('add_file_group.html', facility=facility, parent_group=parent_group)

@app.route('/file/<int:file_id>/move', methods=['GET', 'POST'])
@login_required
def move_file(file_id):
    if current_user.role != 'admin':
        abort(403)  # Запрет доступа, если не админ
    file = Document.query.get_or_404(file_id)
    facilities = Facility.query.order_by(Facility.name).all()
    if request.method == 'POST':
        new_facility_id = request.form.get('facility_id')
        new_group_id = request.form.get('group_id') or None
        new_facility = Facility.query.get_or_404(new_facility_id)
        file.facility_id = new_facility.id
        file.group_id = new_group_id
        db.session.commit()
        flash('Файл успешно перемещен.', 'success')
        return redirect(url_for('facility_detail', facility_id=new_facility.id))
    return render_template('move_file.html', file=file, facilities=facilities)

@app.route('/api/facility/<int:facility_id>/groups')
@login_required
def get_facility_groups(facility_id):
    facility = Facility.query.get_or_404(facility_id)
    groups = [
        {'id': group.id, 'name': group.name}
        for group in facility.file_groups if not group.parent_id
    ]
    return {'groups': groups}

@app.route('/admin/import_users', methods=['POST'])
@admin_required
def import_users_route():
    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        return redirect(url_for('admin_dashboard', message='Загрузите файл Excel (.xlsx)', category='danger'))

    try:
        workbook = load_workbook(file)
        sheet = workbook.active
        count_added = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
            username, first_name, last_name, password, role = row
            if not username or not password:
                continue
            if User.query.filter_by(username=username).first():
                continue
            user = User(
                username=username.strip(),
                first_name=first_name.strip() if first_name else '',
                last_name=last_name.strip() if last_name else '',
                role=role.strip() if role else 'user'
            )
            user.set_password(password.strip())
            db.session.add(user)
            count_added += 1
        db.session.commit()
        resp = redirect(url_for('admin_dashboard'))
        resp.set_cookie("toast_message", f"Импортировано пользователей: {count_added}", max_age=5)
        resp.set_cookie("toast_category", "success", max_age=5)
        return resp
    except Exception as e:
        resp = redirect(url_for('admin_dashboard'))
        resp.set_cookie("toast_message", f"Ошибка при импорте: {str(e)}", max_age=5)
        resp.set_cookie("toast_category", "danger", max_age=5)
        return resp

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)